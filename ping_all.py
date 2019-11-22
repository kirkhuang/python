import openpyxl
import subprocess

# 打开本目录下的excel文件
# wb = openpyxl.load_workbook("D:\\repo\\python\\test.xlsx")
wb = openpyxl.load_workbook("./1.xlsx")
# 百分数显示
# wb.guess_types = True

# ws = wb.active

# 该文件下的所有工作表
sheets = wb.sheetnames

# # 按工作表的名称来指定要操作的工作表
# ws = wb.get_sheet_by_name(sheets[0])

# # 取出字段
# ip = ws["b2"].value

# # ping 这个ip
# str_ip = 'ping ' + ip + ' -l 4000 -w 500 -n 2'
# exit_code = os.system(str_ip)

# # 将结果写入C列
# if exit_code == 0:
#     ws["c2"].value = 'connect success.'
#     print(exit_code)
# else:
#     ws["c2"].value = 'connect failed.'
#     print(exit_code)

# # 保存excel
# wb.save("test.xlsx")


# def pingCHK(ip):
#     str_ip = 'ping ' + ip + ' -l 1000 -w 500 -n 2'
#     exit_code = os.system(str_ip)
#     return exit_code


def pingCHK(ip):
    num = 1
    wait = 1000
    # 这种方式，终端不会显示运行结果
    ping = subprocess.Popen("ping -n {} -w {} {}".format(num, wait, ip), stdout=subprocess.PIPE, stderr=subprocess.PIPE)
    exit_code = ping.wait()
    if exit_code != 0:
        return False
    else:
        return True


for sheet in sheets:
    # print(sheet)
    ws = wb[sheet]
    # ws = wb.get_sheet_by_name(sheet)
    for row in range(1, ws.max_row + 1):
        col = 4
        ip = ws.cell(row, col).value
        if ip:
            print(ip)
            ws.cell(row, col + 1).value = pingCHK(ip)
            print(ws.cell(row, col + 1).value)
        else:
            continue
    for row in range(1, ws.max_row + 1):
        col = 8
        ip = ws.cell(row, col).value
        if ip:
            print(ip)
            ws.cell(row, col + 1).value = pingCHK(ip)
            print(ws.cell(row, col + 1).value)
        else:
            continue
    
wb.save("1.xlsx")
print("检测结束")
