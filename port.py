#!/usr/bin/env python
# coding:utf-8

import socket
import openpyxl


def checkIP(ip):
    if ip:
        try:
            socket.inet_aton(ip)
            return True
        except socket.error:
            return False
    else:
        return False



def telnet(ip, port, timeout=1):
    cs = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    address = (str(ip), int(port))
    status = cs.connect_ex((address))
    cs.settimeout(timeout)
    if status != 0:
        return False
    else:
        return True



# 打开本目录下的excel文件
# wb = openpyxl.load_workbook("D:\\repo\\python\\test.xlsx")
wb = openpyxl.load_workbook("./test.xlsx")

sheets = wb.sheetnames


for sheet in sheets:
    # print(sheet)
    ws = wb[sheet]
    # ws = wb.get_sheet_by_name(sheet)
    for row in range(1, ws.max_row + 1):
        col = 2
        ip = ws.cell(row, col).value
        print(ip)
        if checkIP(ip):
            if telnet(str(ip), 445):
                print(ip)
                ws.cell(row, col + 1).value = 1
                print(ws.cell(row, col + 1).value)
            else:
                continue
        else:
            continue
    
wb.save("test.xlsx")
print("检测结束")
