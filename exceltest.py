import openpyxl

# 打开本目录下的excel文件
# wb = openpyxl.load_workbook("D:\\repo\\python\\test.xlsx")
wb = openpyxl.load_workbook("./test.xlsx")
# 百分数显示
# wb.guess_types = True

ws = wb.active
# 读B2的值
print(ws["b2"].value)

# 读B列的值
print(ws["B"])
for i in ws["B"]:
    print(i.value)


# 打印所有工作簿
print(wb.sheetnames)




for sn in wb.sheetnames:
    print(sn)
